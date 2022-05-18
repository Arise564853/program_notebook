import os
import datetime
import openpyxl as vb


path = 'G:/2020普联检测工地/2020起重报告总汇/'
month = datetime.datetime.today().month
book = vb.Workbook()
worksheet = book.create_sheet(index=0, title='名单')
worksheet.cell(1, 1, value='联系人')
worksheet.cell(1, 2, value='检测日期')
worksheet.cell(1, 3, value='工程名称')
row = 2
for i in ['杭州', '嘉兴']:
    for customer in os.listdir(f'{path}2020{i}工地报告'):
        if os.path.isdir(f'{path}2020{i}工地报告/{customer}'):
            for project in os.listdir(f'{path}2020{i}工地报告/{customer}'):
                try:
                    int(project[:2])
                except:
                    pass
                else:
                    if int(project[:2]) == month:
                        worksheet.cell(row, 1, value=customer)
                        worksheet.cell(row, 2, value=project[:4])
                        worksheet.cell(row, 3, value=project[4:])
                        row += 1
book.save(f'2021年{month}月复检到期.xlsx')

